import os
import glob
import requests
import shutil
import uuid
import polars as pl
import connectorx as cx
import fastexcel
import tempfile
import time
from openpyxl import load_workbook
from typing import List, Literal, Optional, Any, Dict, cast

STAGING_DIR_NAME = "pyquery_staging"


def get_staging_dir() -> str:
    """Get or create the centralized staging directory."""
    temp_dir = tempfile.gettempdir()
    staging_path = os.path.join(temp_dir, STAGING_DIR_NAME)
    os.makedirs(staging_path, exist_ok=True)
    return staging_path


def cleanup_staging_files(max_age_hours: int = 24):
    """Clean up old files from the staging directory."""
    try:
        staging_dir = get_staging_dir()
        now = time.time()
        cutoff = now - (max_age_hours * 3600)

        if os.path.exists(staging_dir):
            for filename in os.listdir(staging_dir):
                file_path = os.path.join(staging_dir, filename)
                try:
                    if os.path.isfile(file_path):
                        if os.path.getmtime(file_path) < cutoff:
                            os.remove(file_path)
                            # print(f"Cleaned up stale staging file: {filename}") # Reduce noise
                except Exception as e:
                    print(f"Failed to delete {filename}: {e}")
    except Exception as e:
        print(f"Cleanup Error: {e}")


def get_files_from_path(path_str: str) -> List[str]:
    if not path_str:
        return []
    if "*" in path_str:
        return glob.glob(path_str)
    return [path_str] if os.path.exists(path_str) else []


def get_excel_sheet_names(file_path: str) -> List[str]:
    """
    Efficiently retrieve sheet names from an Excel file using fastexcel.
    Fallback to 'Sheet1' if any error occurs.
    Handles globs and directories by inspecting the first matching file.
    """
    try:
        # Resolve path (handle globs, dirs)
        files = get_files_from_path(file_path)
        if not files:
            return ["Sheet1"]

        target_file = files[0]

        ext = os.path.splitext(target_file)[1].lower()
        if ext not in [".xlsx", ".xls", ".xlsm"]:
            return ["Sheet1"]

        try:
            excel = fastexcel.read_excel(target_file)
            return excel.sheet_names
        except Exception:
             # Fallback
            try:
                wb = load_workbook(
                    target_file, read_only=True, keep_links=False)
                return wb.sheetnames
            except:
                return ["Sheet1"]
    except Exception:
        return ["Sheet1"]


def load_lazy_frame(files: List[str], sheet_name: str = "Sheet1") -> Optional[pl.LazyFrame]:
    if not files:
        return None

    # OPTIMIZATION: Try Bulk Scan for homogeneous files
    # Polars supports list of files for scan_csv, scan_parquet, scan_ipc, scan_ndjson
    # This is much faster than iterative concat.
    exts = {os.path.splitext(f)[1].lower() for f in files}
    if len(exts) == 1:
        ext = list(exts)[0]
        try:
            if ext == ".csv":
                return pl.scan_csv(files, infer_schema_length=0)
            elif ext == ".parquet":
                return pl.scan_parquet(files)
            elif ext in [".arrow", ".ipc", ".feather"]:
                return pl.scan_ipc(files)
            elif ext == ".json":
                return pl.scan_ndjson(files, infer_schema_length=0)
        except Exception as e:
            print(f"Bulk scan error, falling back to iterative: {e}")

    # Fallback: Iterative (Mixed types or Excel)
    lfs = []
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        try:
            if ext == ".csv":
                lfs.append(pl.scan_csv(f, infer_schema_length=0))
            elif ext == ".parquet":
                lfs.append(pl.scan_parquet(f))
            elif ext in [".arrow", ".ipc", ".feather"]:
                lfs.append(pl.scan_ipc(f))
            elif ext in [".xlsx", ".xls"]:
                # Dump to Parquet
                # Excel is eager-only.
                try:
                    df = pl.read_excel(f, sheet_name=sheet_name,
                                       infer_schema_length=0)

                    # STAGING: Use file's directory instead of script root
                    # Prevents cluttering the app directory
                    base_dir = os.path.dirname(os.path.abspath(f))
                    staging_dir = os.path.join(base_dir, ".staging")
                    os.makedirs(staging_dir, exist_ok=True)

                    stage_filename = f"{os.path.splitext(os.path.basename(f))[0]}_{uuid.uuid4().hex[:8]}.parquet"
                    stage_path = os.path.join(staging_dir, stage_filename)

                    df.write_parquet(stage_path)
                    del df

                    lfs.append(pl.scan_parquet(stage_path))

                except Exception as ex:
                    print(f"Excel Load Error {f}: {ex}")

            elif ext == ".json":
                lfs.append(pl.scan_ndjson(f, infer_schema_length=0))
        except Exception as e:
            print(f"Error loading {f}: {e}")

    if not lfs:
        return None

    combined = lfs[0]

    # DIAGONAL STRATEGY:
    # Handles schema evolution (new columns in later files filled with nulls).
    # Graceful handling of missing columns.
    if len(lfs) > 1:
        combined = pl.concat(lfs, how="diagonal")

    return combined


def load_from_sql(connection_string: str, query: str) -> Optional[pl.LazyFrame]:
    try:
        # connectorx returns eager Arrow/DataFrame, we make it lazy
        # This is strictly backend logic (IO)
        df_arrow = cx.read_sql(connection_string, query, return_type="arrow")
        df = pl.from_arrow(df_arrow)

        # Ensure it's a DataFrame before calling lazy
        if isinstance(df, pl.Series):
            df = df.to_frame()

        return df.lazy()
    except Exception as e:
        print(f"SQL Error: {e}")
        return None


def load_from_api(url: str) -> Optional[pl.LazyFrame]:
    try:
        # Enterprise Staged Loading: Stream to disk first
        staging_dir = os.path.join(os.getcwd(), ".staging")
        os.makedirs(staging_dir, exist_ok=True)

        file_name = f"api_dump_{uuid.uuid4()}.json"
        file_path = os.path.join(staging_dir, file_name)

        # Stream download (low memory usage)
        with requests.get(url, stream=True) as r:
            r.raise_for_status()
            with open(file_path, 'wb') as f:
                shutil.copyfileobj(r.raw, f)

        # Return LazyFrame from disk
        return pl.read_json(file_path).lazy()
    except Exception as e:
        print(f"API Error: {e}")
        return None


def export_worker(lazy_frame: pl.LazyFrame, params: Any, fmt: str, result_container: Dict[str, Any]):
    try:
        # Extract path safely from params (Dict or Pydantic)
        path = params.get('path') if isinstance(
            params, dict) else getattr(params, 'path', None)
        if not path:
            raise ValueError("Output path not specified")

        # OPTIMIZATION: Use Streaming Sinks where possible
        if fmt == "CSV":
            # sink_csv is streaming
            lazy_frame.sink_csv(path)

        elif fmt == "Parquet":
            compression = params.get('compression', 'snappy') if isinstance(
                params, dict) else getattr(params, 'compression', 'snappy')
            valid_compression = cast(
                Literal['snappy', 'zstd', 'gzip', 'lz4', 'uncompressed', 'brotli'], compression)
            # sink_parquet is streaming
            lazy_frame.sink_parquet(path, compression=valid_compression)

        elif fmt == "IPC":
            compression = params.get('compression', 'uncompressed') if isinstance(
                params, dict) else getattr(params, 'compression', 'uncompressed')
            valid_compression = cast(
                Literal['uncompressed', 'lz4', 'zstd'], compression)
            # sink_ipc is streaming
            lazy_frame.sink_ipc(path, compression=valid_compression)

        elif fmt == "NDJSON":
            # sink_ndjson is streaming
            lazy_frame.sink_ndjson(path)

        elif fmt == "Excel":
            # Native Polars (Fast Eager Write)
            df = lazy_frame.collect()
            df.write_excel(path)

        elif fmt == "JSON":
            # Native Polars (Fast Eager Write)
            df = lazy_frame.collect()
            df.write_json(path)

        elif fmt == "SQLite":
            # SQLite Export (Eager)
            table = params.get('table', 'data') if isinstance(
                params, dict) else getattr(params, 'table', 'data')
            if_exists = params.get('if_exists', 'replace') if isinstance(
                params, dict) else getattr(params, 'if_exists', 'replace')
            valid_if_exists = cast(
                Literal['fail', 'replace', 'append'], if_exists)

            df = lazy_frame.collect()
            # Construct connection string
            # write_database supports "sqlite:///path.db"
            uri = f"sqlite:///{path}"
            df.write_database(table_name=table, connection=uri,
                              if_table_exists=valid_if_exists, engine="sqlalchemy")

        result_container['status'] = "Done"
    except Exception as e:
        result_container['status'] = f"Error: {e}"
