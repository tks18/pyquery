from typing import List, Dict, Any

import os
import fastexcel
from openpyxl import load_workbook

from pyquery_polars.backend.io.helpers.filters import FilterEngine


class ExcelEngine:
    """
    Utilities for working with Excel files
    """

    @classmethod
    def get_excel_metadata(cls, file_path: str) -> Dict[str, Any]:
        """
        Single-pass Excel metadata extraction.
        Returns sheet names, table names, and basic file info.
        Caches results to avoid multiple file reads.

        Returns:
            Dict with keys: 'sheets' (List[str]), 'tables' (List[str]), 'valid' (bool)
        """
        metadata = {
            'sheets': ["Sheet1"],
            'tables': [],
            'valid': False
        }

        try:
            # Resolve path (handle globs, dirs)
            files = FilterEngine.resolve_file_paths(file_path)
            if not files:
                return metadata

            target_file = files[0]
            ext = os.path.splitext(target_file)[1].lower()

            if ext not in [".xlsx", ".xls", ".xlsm", ".xlsb"]:
                return metadata

            # Single fastexcel reader instantiation
            try:
                reader = fastexcel.read_excel(target_file)
                metadata['sheets'] = reader.sheet_names if reader.sheet_names else [
                    "Sheet1"]

                # Get tables (only for formats that support it)
                if ext in [".xlsx", ".xlsm", ".xlsb"]:
                    try:
                        metadata['tables'] = sorted(reader.table_names())
                    except Exception:
                        metadata['tables'] = []

                metadata['valid'] = True
                return metadata

            except Exception as e:
                # Fallback to openpyxl for sheets only
                try:
                    wb = load_workbook(
                        target_file, read_only=True, keep_links=False)
                    metadata['sheets'] = wb.sheetnames if wb.sheetnames else [
                        "Sheet1"]
                    metadata['valid'] = True
                    wb.close()
                    return metadata
                except:
                    return metadata

        except Exception:
            return metadata

    @classmethod
    def get_excel_sheet_names(cls, file_path: str) -> List[str]:
        """
        Efficiently retrieve sheet names from an Excel file.
        Uses cached metadata extraction to avoid redundant reads.
        Fallback to 'Sheet1' if any error occurs.
        """
        metadata = cls.get_excel_metadata(file_path)
        return metadata['sheets']

    @classmethod
    def get_excel_table_names(cls, file_path: str) -> List[str]:
        """
        Retrieve defined Table names from an Excel file.
        Uses cached metadata extraction to avoid redundant reads.
        """
        metadata = cls.get_excel_metadata(file_path)
        return metadata['tables']
