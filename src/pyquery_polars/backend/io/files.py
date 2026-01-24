import re
import os
import glob
import requests
import shutil
import uuid
import polars as pl
import connectorx as cx
import chardet
import fastexcel
import tempfile
import time
import copy
import io
import codecs
import gc
from itertools import islice
from openpyxl import load_workbook
from typing import List, Literal, Optional, Any, Dict, cast, Iterator, Union
import fnmatch

from ...core.io_params import FileFilter, ItemFilter, FilterType

STAGING_DIR_NAME = "pyquery_staging"


def get_staging_dir() -> str:
    """Get or create the centralized staging directory."""
    temp_dir = tempfile.gettempdir()
    staging_path = os.path.join(temp_dir, STAGING_DIR_NAME)
    os.makedirs(staging_path, exist_ok=True)
    return staging_path


def create_unique_staging_folder(base_name: str) -> str:
    """
    Create a unique subfolder in the staging directory.
    Format: timestamp_uuid_basename
    """
    staging_root = get_staging_dir()
    
    # Generate unique identifier components
    ts = int(time.time())
    unique_id = uuid.uuid4().hex[:8]
    
    # Sanitize base name
    safe_name = re.sub(r'[^a-zA-Z0-9_.-]', '_', base_name)
    
    folder_name = f"{ts}_{unique_id}_{safe_name}"
    folder_path = os.path.join(staging_root, folder_name)
    
    os.makedirs(folder_path, exist_ok=True)
    return folder_path


def cleanup_staging_files(max_age_hours: int = 24):
    """Clean up old files from the staging directory."""
    try:
        staging_dir = get_staging_dir()
        now = time.time()
        cutoff = now - (max_age_hours * 3600)

        if os.path.exists(staging_dir):
            for filename in os.listdir(staging_dir):
                file_path = os.path.join(staging_dir, filename)
                try:
                    if os.path.isfile(file_path):
                        if os.path.getmtime(file_path) < cutoff:
                            os.remove(file_path)
                    elif os.path.isdir(file_path):
                        # Clean up stale directories
                        if os.path.getmtime(file_path) < cutoff:
                            shutil.rmtree(file_path)
                except Exception as e:
                    pass
    except Exception as e:
        pass


def cleanup_staging_file(file_path: str):
    """Remove a specific staging file immediately."""
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception:
        pass


def resolve_file_paths(base_path: str, filters: Optional[List[FileFilter]] = None, limit: Optional[int] = None) -> List[str]:
    """
    Resolve a base path and optional filters into a list of file paths.

    Uses a streaming approach to efficiently handle large directories.
    Optimizes simple filters into glob patterns where possible (Partial Globbing).
    Falls back to Python-side filtering for complex requirements.

    Args:
        base_path: Target path (file, directory, or glob pattern).
        filters: Optional list of filters to apply.
        limit: Optional maximum number of files to return (for previews).
    """
    if not base_path:
        return []

    # Scenario 1: No Filters
    # If a specific path or valid glob is provided without extra filters,
    # we return it directly to let the optimized Polars reader handle scanning.
    if not filters:
        if os.path.isfile(base_path):
            return [base_path]
        if "*" in base_path:
            # standard behavior for 'resolve' implies returning the list.
            if limit:
                return list(islice(glob.iglob(base_path, recursive="**" in base_path), limit))
            return glob.glob(base_path, recursive="**" in base_path)

        if os.path.isdir(base_path):
            # Return directory as-is for Polars to scan/hive-partition auto-detect
            return [base_path]

        return []

    # Scenario 2: Filters Present
    # We must scan and filter files manually (or partially optimized).

    # Attempt to narrow the search space using the most restrictive filter (Partial Globbing)
    optimized_glob = _optimize_filters_to_glob(base_path, filters)

    candidates_iter: Iterator[str]

    if optimized_glob:
        # Use the optimized glob as the primary candidate source
        candidates_iter = glob.iglob(
            optimized_glob, recursive="**" in optimized_glob)
    else:
        # Fallback: Determine candidate source based on base_path type
        if "*" in base_path:
            candidates_iter = glob.iglob(
                base_path, recursive="**" in base_path)
        elif os.path.isdir(base_path):
            # Generator for recursive directory scan
            candidates_iter = _recursive_dir_walker(base_path)
        elif os.path.isfile(base_path):
            candidates_iter = iter([base_path])
        else:
            return []

    # Apply remaining filters in a streaming fashion
    return _apply_param_filters(candidates_iter, filters, limit)


def _recursive_dir_walker(path: str) -> Iterator[str]:
    """Yields all file paths recursively from a directory."""
    for dp, dn, filenames in os.walk(path):
        for f in filenames:
            yield os.path.join(dp, f)


def _optimize_filters_to_glob(base_path: str, filters: List[FileFilter]) -> Optional[str]:
    """
    Selects the best available filter to create a narrowing glob pattern.
    Priority: EXACT > GLOB > CONTAINS (filename target only).
    """
    # Globbing is only applicable if we are starting from a directory
    if not os.path.isdir(base_path):
        return None

    # Priority 1: Exact Filename Match
    for f in filters:
        if f.target == "filename" and f.type == FilterType.EXACT:
            return os.path.join(base_path, f.value)

    # Priority 2: User-provided Glob
    for f in filters:
        if f.target == "filename" and f.type == FilterType.GLOB:
            return os.path.join(base_path, f.value)

    # Priority 3: Contains (Substring)
    for f in filters:
        if f.target == "filename" and f.type == FilterType.CONTAINS:
            return os.path.join(base_path, "**", f"*{f.value}*")

    return None


def _check_filter_match(path: str, f: FileFilter) -> bool:
    """Evaluates if a file path satisfies a single filter."""
    val = f.value

    # Resolve target
    if f.target == "path":
        check_val = path
    else:
        check_val = os.path.basename(path)

    # Standardize case for case-insensitive comparisons
    # EXACT is the only strictly case-sensitive mode
    check_lower = check_val.lower()
    val_lower = val.lower()

    if f.type == FilterType.EXACT:
        return val == check_val

    if f.type == FilterType.IS_NOT:
        return val != check_val

    if f.type == FilterType.CONTAINS:
        return val_lower in check_lower

    if f.type == FilterType.NOT_CONTAINS:
        return val_lower not in check_lower

    if f.type == FilterType.GLOB:
        return fnmatch.fnmatch(check_lower, val_lower)

    if f.type == FilterType.REGEX:
        try:
            return bool(re.search(val, check_val, re.IGNORECASE))
        except re.error:
            return False

    return False


def _check_item_match(name: str, f: ItemFilter) -> bool:
    """Evaluates if a sheet name satisfies a filter."""
    val = f.value
    check_val = name

    # Standardize case for case-insensitive comparisons
    # EXACT is the only strictly case-sensitive mode
    check_lower = check_val.lower()
    val_lower = val.lower()

    if f.type == FilterType.EXACT:
        return val == check_val

    if f.type == FilterType.IS_NOT:
        return val != check_val

    if f.type == FilterType.CONTAINS:
        return val_lower in check_lower

    if f.type == FilterType.NOT_CONTAINS:
        return val_lower not in check_lower

    if f.type == FilterType.GLOB:
        return fnmatch.fnmatch(check_lower, val_lower)

    if f.type == FilterType.REGEX:
        try:
            return bool(re.search(val, check_val, re.IGNORECASE))
        except re.error:
            return False

    return False


def _apply_param_filters(files: Iterator[str], filters: List[FileFilter], limit: Optional[int] = None) -> List[str]:
    """
    Consumes the file iterator, applies filters, and returns a list up to the limit.
    """
    kept = []

    for path in files:
        # Check limit before processing
        if limit is not None and len(kept) >= limit:
            break

        # Verify all filters match
        if all(_check_filter_match(path, f) for f in filters):
            kept.append(path)

    return kept


def get_files_from_path(path_str: str) -> List[str]:
    # Backward compatibility wrapper
    return resolve_file_paths(path_str)


def _get_excel_metadata(file_path: str) -> Dict[str, Any]:
    """
    Single-pass Excel metadata extraction.
    Returns sheet names, table names, and basic file info.
    Caches results to avoid multiple file reads.

    Returns:
        Dict with keys: 'sheets' (List[str]), 'tables' (List[str]), 'valid' (bool)
    """
    metadata = {
        'sheets': ["Sheet1"],
        'tables': [],
        'valid': False
    }

    try:
        # Resolve path (handle globs, dirs)
        files = get_files_from_path(file_path)
        if not files:
            return metadata

        target_file = files[0]
        ext = os.path.splitext(target_file)[1].lower()

        if ext not in [".xlsx", ".xls", ".xlsm", ".xlsb"]:
            return metadata

        # Single fastexcel reader instantiation
        try:
            reader = fastexcel.read_excel(target_file)
            metadata['sheets'] = reader.sheet_names if reader.sheet_names else [
                "Sheet1"]

            # Get tables (only for formats that support it)
            if ext in [".xlsx", ".xlsm", ".xlsb"]:
                try:
                    metadata['tables'] = sorted(reader.table_names())
                except Exception:
                    metadata['tables'] = []

            metadata['valid'] = True
            return metadata

        except Exception as e:
            # Fallback to openpyxl for sheets only
            try:
                wb = load_workbook(
                    target_file, read_only=True, keep_links=False)
                metadata['sheets'] = wb.sheetnames if wb.sheetnames else [
                    "Sheet1"]
                metadata['valid'] = True
                wb.close()
                return metadata
            except:
                return metadata

    except Exception:
        return metadata


def get_excel_sheet_names(file_path: str) -> List[str]:
    """
    Efficiently retrieve sheet names from an Excel file.
    Uses cached metadata extraction to avoid redundant reads.
    Fallback to 'Sheet1' if any error occurs.
    """
    metadata = _get_excel_metadata(file_path)
    return metadata['sheets']


def get_excel_table_names(file_path: str) -> List[str]:
    """
    Retrieve defined Table names from an Excel file.
    Uses cached metadata extraction to avoid redundant reads.
    """
    metadata = _get_excel_metadata(file_path)
    return metadata['tables']


def clean_header_name(col: str) -> str:
    """Normalize column name by replacing whitespace with single spaces and stripping."""
    return " ".join(col.strip().split())


def detect_encoding(file_path: str, limit_bytes: int = 200_000) -> str:
    """
    Robustly detect file encoding using streaming analysis (UniversalDetector).
    Scans up to `limit_bytes` (default 200KB) or until high confidence is reached.
    """
    try:
        from chardet.universaldetector import UniversalDetector
        detector = UniversalDetector()

        # Read in binary mode, 16KB chunks
        chunk_size = 16 * 1024
        processed_bytes = 0

        with open(file_path, 'rb') as f:
            while processed_bytes < limit_bytes:
                chunk = f.read(chunk_size)
                if not chunk:
                    break

                detector.feed(chunk)
                processed_bytes += len(chunk)

                if detector.done:
                    break

        detector.close()
        result = detector.result

        encoding = result['encoding']
        confidence = result['confidence']

        # Confidence Threshold
        if not encoding or (confidence and confidence < 0.6):
            return 'utf-8'

        # Normalize typical compatible encodings
        if encoding.lower() in ['ascii', 'utf-8-sig']:
            return 'utf-8'

        return encoding

    except Exception:
        # Fallback to UTF-8 on any error
        return 'utf-8'


def batch_detect_encodings(files: List[str]) -> Dict[str, str]:
    """
    Detect encodings for a list of files.
    Returns a dictionary mapping file path to detected encoding.
    Only returns entries where encoding is NOT utf8 (or ascii).
    """
    results = {}
    for f in files:
        # Skip if not a text file (simplified check)
        ext = os.path.splitext(f)[1].lower()
        if ext not in [".csv", ".txt", ".json", ".ndjson"]:
            continue

        enc = detect_encoding(f)
        # Normalize: ascii is compatible with utf8
        if enc.lower() not in ['utf8', 'utf-8', 'ascii']:
            results[f] = enc
    return results


def convert_file_to_utf8(file_path: str, source_encoding: str) -> str:
    """
    Convert a file from source_encoding to UTF-8 using robust streaming.
    Features:
    - Normalizes newlines to '\\n' (critical for Polars CSV parser robustness)
    - Removes NULL bytes (\\x00)
    - Uses 'replace' error handler for garbage characters
    - Streams in 4MB chunks for low RAM usage
    """
    # 1. Validate Encoding
    try:
        if not source_encoding:
            source_encoding = "utf-8"
        codecs.lookup(source_encoding)
    except LookupError:
        print(
            f"Warning: Encoding '{source_encoding}' not found, falling back to 'utf-8'.")
        source_encoding = "utf-8"

    new_path = None
    try:
        staging_dir = get_staging_dir()
        filename = os.path.basename(file_path)

        # Sanitize filename to prevent issues
        safe_name = re.sub(r'[^a-zA-Z0-9_.-]', '_', filename)
        new_filename = f"utf8_{uuid.uuid4().hex[:6]}_{safe_name}"
        new_path = os.path.join(staging_dir, new_filename)

        # 4MB Chunk Size for better IO throughput
        chunk_size = 4 * 1024 * 1024

        # Open Source: newline=None enables Universal Newlines (normalizes \r\n to \n)
        # Open Target: newline='\n' ensures we write clean Unix-style output
        with io.open(file_path, 'r', encoding=source_encoding, errors='replace', newline=None) as source_f:
            with io.open(new_path, 'w', encoding='utf-8', newline='\n') as target_f:
                while True:
                    chunk = source_f.read(chunk_size)
                    if not chunk:
                        break

                    # Robustness: Remove NULL bytes which confuse C-parsers
                    if '\0' in chunk:
                        chunk = chunk.replace('\0', '')

                    target_f.write(chunk)

        return new_path

    except Exception as e:
        print(f"Failed to convert {file_path}: {e}")
        # Attempt cleanup of partial file
        # Check if new_path was ever assigned (it is local to try block, but accessible in except if assigned)
        # But to be safe against 'possibly unbound' (if exception happened before assignment), we check
        if 'new_path' in locals() and new_path and os.path.exists(new_path):
            try:
                os.remove(new_path)
            except:
                pass
        raise e


def load_lazy_frame(files: List[str], sheet_name: Optional[Union[str, List[str]]] = "Sheet1", sheet_filters: Optional[List[ItemFilter]] = None, table_name: Optional[Union[str, List[str]]] = None, table_filters: Optional[List[ItemFilter]] = None, process_individual: bool = False, include_source_info: bool = False, clean_headers: bool = False) -> Optional[tuple]:
    """
    Load files into LazyFrame(s).

    Returns:
        Tuple of (Union[LazyFrame, List[LazyFrame]], metadata_dict) or None
    """
    if not files:
        return None

    # Determine file format
    exts = {os.path.splitext(f)[1].lower() for f in files}
    ext = list(exts)[0] if len(exts) == 1 else ".mixed"

    # OPTIMIZATION: Try Bulk Scan for homogeneous files
    # Only if NOT processing individual, NOT including source info, AND NOT cleaning headers
    if len(exts) == 1 and not process_individual and not include_source_info and not clean_headers:
        try:
            if ext == ".csv":
                # Strict UTF-8: We assume files are UTF-8 validated by the frontend/pre-check.
                # If they are not, this will likely fail, which satisfies the "reject" requirement.
                lf = pl.scan_csv(files, infer_schema_length=0, encoding="utf8")
                metadata = {
                    "input_type": "folder" if len(files) > 1 else "file",
                    "input_format": ext,
                    "file_list": files,
                    "file_count": len(files)
                }
                return lf, metadata
            elif ext == ".parquet":
                lf = pl.scan_parquet(files)
                metadata = {
                    "input_type": "folder" if len(files) > 1 else "file",
                    "input_format": ext,
                    "file_list": files,
                    "file_count": len(files)
                }
                return lf, metadata
            elif ext in [".arrow", ".ipc", ".feather"]:
                lf = pl.scan_ipc(files)
                metadata = {
                    "input_type": "folder" if len(files) > 1 else "file",
                    "input_format": ext,
                    "file_list": files,
                    "file_count": len(files)
                }
                return lf, metadata
            elif ext == ".json":
                lf = pl.scan_ndjson(files, infer_schema_length=0)
                metadata = {
                    "input_type": "folder" if len(files) > 1 else "file",
                    "input_format": ext,
                    "file_list": files,
                    "file_count": len(files)
                }
                return lf, metadata
        except Exception as e:
            print(f"Bulk scan error, falling back to iterative: {e}")

    # Fallback: Iterative
    lfs = []
    for f in files:
        file_ext = os.path.splitext(f)[1].lower()
        try:
            current_lf = None
            if file_ext == ".csv":
                # Strict UTF-8
                current_lf = pl.scan_csv(
                    f, infer_schema_length=0, encoding="utf8")
            elif file_ext == ".parquet":
                current_lf = pl.scan_parquet(f)
            elif file_ext in [".arrow", ".ipc", ".feather"]:
                current_lf = pl.scan_ipc(f)
            elif file_ext in [".xlsx", ".xls", ".xlsm", ".xlsb"]:
                try:
                    staging_path = get_staging_dir()

                    # OPTIMIZATION: Single metadata extraction per file
                    excel_meta = _get_excel_metadata(f)

                    # Determine what to load
                    # Priority: Table Name(s) > Sheet Name(s) > Default Sheet1

                    # Normalize inputs to lists
                    target_tables = []

                    if table_name == "__ALL_TABLES__" or table_name == ["__ALL_TABLES__"]:
                        target_tables = excel_meta['tables']
                    elif table_name:
                        if isinstance(table_name, list):
                            target_tables = table_name
                        else:
                            target_tables = [table_name]

                    target_sheets = []

                    if not target_tables:
                        if table_filters:
                            # DYNAMIC TABLE SELECTION
                            all_tables = excel_meta['tables']
                            # Reuse _check_item_match (generic enough)
                            target_tables = [t for t in all_tables if all(
                                _check_item_match(t, tf) for tf in table_filters)]

                        elif sheet_filters is not None:
                            # DYNAMIC SHEET SELECTION
                            all_sheets = excel_meta['sheets']
                            # Apply filters
                            target_sheets = [s for s in all_sheets if all(
                                _check_item_match(s, sf) for sf in sheet_filters)]

                        elif sheet_name == "__ALL_SHEETS__" or sheet_name == ["__ALL_SHEETS__"]:
                            target_sheets = excel_meta['sheets']

                        elif isinstance(sheet_name, list):
                            target_sheets = sheet_name

                        else:
                            # Single sheet string or None -> Default
                            if sheet_name:
                                target_sheets = [sheet_name]
                            else:
                                # Use first sheet from metadata
                                if excel_meta['sheets']:
                                    target_sheets = [excel_meta['sheets'][0]]
                                else:
                                    target_sheets = ["Sheet1"]

                    # 1. LOAD TABLES (via Polars read_excel -> Parquet)
                    if target_tables:
                        for t_name in target_tables:
                            try:
                                # Polars read_excel with table_name
                                df = pl.read_excel(
                                    f, table_name=t_name, engine="calamine", infer_schema_length=0)

                                if clean_headers:
                                    new_cols = {c: clean_header_name(
                                        c) for c in df.columns}
                                    df = df.rename(new_cols)

                                if include_source_info:
                                    df = df.with_columns([
                                        pl.lit(os.path.abspath(f)).alias(
                                            "__pyquery_source_path__"),
                                        pl.lit(f"{os.path.basename(f)}[table][{t_name}]").alias(
                                            "__pyquery_source_name__"),
                                        pl.lit(file_ext).alias(
                                            "__pyquery_source_ext__")
                                    ])

                                # Write to Staging
                                out_name = f"staged_{uuid.uuid4().hex[:8]}_{t_name}.parquet"
                                out_path = os.path.join(staging_path, out_name)
                                df.write_parquet(out_path)

                                # MEMORY CLEANUP: Explicit deletion
                                del df

                                # Append LazyFrame reference
                                lfs.append(pl.scan_parquet(out_path))

                            except Exception as e:
                                print(f"Failed to load table {t_name}: {e}")

                    # 2. LOAD SHEETS (via Polars read_excel -> Parquet)
                    if target_sheets:
                        for s_name in target_sheets:
                            try:
                                # pl.read_excel is eager
                                df = pl.read_excel(
                                    f, sheet_name=s_name, engine="calamine", infer_schema_length=0)

                                if clean_headers:
                                    new_cols = {c: clean_header_name(
                                        c) for c in df.columns}
                                    df = df.rename(new_cols)

                                if include_source_info:
                                    df = df.with_columns([
                                        pl.lit(os.path.abspath(f)).alias(
                                            "__pyquery_source_path__"),
                                        pl.lit(f"{os.path.basename(f)}[sheet][{s_name}]").alias(
                                            "__pyquery_source_name__"),
                                        pl.lit(file_ext).alias(
                                            "__pyquery_source_ext__")
                                    ])

                                # Write to Staging
                                out_name = f"staged_{uuid.uuid4().hex[:8]}_{s_name}.parquet"
                                out_path = os.path.join(staging_path, out_name)
                                df.write_parquet(out_path)

                                # MEMORY CLEANUP: Explicit deletion
                                del df

                                # Append LazyFrame reference
                                lfs.append(pl.scan_parquet(out_path))

                            except Exception as e:
                                print(f"Failed to load sheet {s_name}: {e}")

                    # Ensure common block is skipped
                    current_lf = None

                except Exception as ex:
                    print(f"Excel Load Error {f}: {ex}")
                    # Ensure common block is skipped on error too
                    current_lf = None

            elif file_ext == ".json":
                current_lf = pl.scan_ndjson(f, infer_schema_length=0)

            # --- POST-SCAN PROCESSING (Common) ---
            if current_lf is not None:
                # 1. Clean Headers (Lazy Rename)
                if clean_headers and file_ext != ".xlsx" and file_ext != ".xls":
                    # We need the schema to rename. collect_schema() is fast.
                    try:
                        base_cols = current_lf.collect_schema().names()
                        rename_map = {c: clean_header_name(
                            c) for c in base_cols}
                        current_lf = current_lf.rename(rename_map)
                    except Exception as e:
                        print(f"Header cleaning failed for {f}: {e}")

                # 2. Source Info
                if include_source_info:
                    abs_path = os.path.abspath(f)
                    name = os.path.basename(f)
                    ext_val = os.path.splitext(f)[1]

                    current_lf = current_lf.with_columns([
                        pl.lit(abs_path).alias("__pyquery_source_path__"),
                        pl.lit(name).alias("__pyquery_source_name__"),
                        pl.lit(ext_val).alias("__pyquery_source_ext__")
                    ])

                lfs.append(current_lf)

        except Exception as e:
            print(f"Error loading {f}: {e}")

    # MEMORY OPTIMIZATION: Suggest garbage collection after batch processing
    if len(files) > 5:  # Only for batch operations
        gc.collect()

    if not lfs:
        return None

    # Build metadata
    metadata = {
        "input_type": "folder" if len(files) > 1 else "file",
        "input_format": ext,
        "file_list": files,
        "file_count": len(files),
        "process_individual": process_individual,
        "source_info_included": include_source_info,
        "clean_headers": clean_headers
    }

    # Decision: Return list or concatenated
    if process_individual and len(lfs) > 1:
        return lfs, metadata
    else:
        combined = lfs[0]
        if len(lfs) > 1:
            combined = pl.concat(lfs, how="diagonal")
        return combined, metadata


def load_from_sql(connection_string: str, query: str) -> Optional[pl.LazyFrame]:
    try:
        # connectorx returns eager Arrow/DataFrame, we make it lazy
        # This is strictly backend logic (IO)
        df_arrow = cx.read_sql(connection_string, query, return_type="arrow")
        df = pl.from_arrow(df_arrow)

        # Ensure it's a DataFrame before calling lazy
        if isinstance(df, pl.Series):
            df = df.to_frame()

        return df.lazy()
    except Exception as e:
        print(f"SQL Error: {e}")
        return None


def load_from_api(url: str) -> Optional[pl.LazyFrame]:
    try:
        # Enterprise Staged Loading: Stream to disk first
        staging_dir = get_staging_dir()

        file_name = f"api_dump_{uuid.uuid4()}.json"
        file_path = os.path.join(staging_dir, file_name)

        # Stream download (low memory usage)
        with requests.get(url, stream=True) as r:
            r.raise_for_status()
            with open(file_path, 'wb') as f:
                shutil.copyfileobj(r.raw, f)

        # Return LazyFrame from disk
        return pl.read_json(file_path).lazy()
    except Exception as e:
        print(f"API Error: {e}")
        return None


def export_worker(lazy_frame: Union[pl.LazyFrame, List[pl.LazyFrame]], params: Any, fmt: str, result_container: Dict[str, Any]):
    try:
        # Extract path safely from params (Dict or Pydantic)
        base_path = params.get('path') if isinstance(
            params, dict) else getattr(params, 'path', None)
        if not base_path:
            raise ValueError("Output path not specified")

        # Ensure directory exists
        os.makedirs(os.path.dirname(base_path), exist_ok=True)

        # --- RECURSIVE HANDLE LIST (Individual Files) ---
        if isinstance(lazy_frame, list):
            # Iterate and export each

            # Decompose path: "folder/data.csv" -> "folder/data" + ".csv"
            p_root, p_ext = os.path.splitext(base_path)

            total_files = len(lazy_frame)
            all_file_details = []

            for i, lf in enumerate(lazy_frame):
                # Construct sub-path
                sub_path = f"{p_root}_{i}{p_ext}"

                # Clone params to override path
                if isinstance(params, dict):
                    sub_params = params.copy()
                    sub_params['path'] = sub_path
                else:
                    # Pydantic copy
                    sub_params = params.model_copy()
                    # Safe set (handling frozen?) usually Pydantic models aren't frozen unless specified
                    setattr(sub_params, 'path', sub_path)

                # Recursive call (safe because we pass single LF)
                # We use a dummy container to catch potential errors per file
                dummy_res = {}
                export_worker(lf, sub_params, fmt, dummy_res)

                if str(dummy_res.get('status', '')).startswith("Error"):
                    raise RuntimeError(
                        f"File {i} failed: {dummy_res['status']}")

                if 'file_details' in dummy_res and dummy_res['file_details']:
                    all_file_details.extend(dummy_res['file_details'])

            result_container['status'] = "Done"
            result_container['size_str'] = f"{total_files} files"
            result_container['file_details'] = all_file_details
            return

        # --- SINGLE FILE EXPORT ---
        # Re-assign for clarity
        path = base_path

        # OPTIMIZATION: Use Streaming Sinks where possible
        if fmt == "CSV":
            # sink_csv is streaming
            lazy_frame.sink_csv(path)

        elif fmt == "Parquet":
            compression = params.get('compression', 'snappy') if isinstance(
                params, dict) else getattr(params, 'compression', 'snappy')
            valid_compression = cast(
                Literal['snappy', 'zstd', 'gzip', 'lz4', 'uncompressed', 'brotli'], compression)
            # sink_parquet is streaming
            lazy_frame.sink_parquet(path, compression=valid_compression)

        elif fmt == "IPC":
            compression = params.get('compression', 'uncompressed') if isinstance(
                params, dict) else getattr(params, 'compression', 'uncompressed')
            valid_compression = cast(
                Literal['uncompressed', 'lz4', 'zstd'], compression)
            # sink_ipc is streaming
            lazy_frame.sink_ipc(path, compression=valid_compression)

        elif fmt == "NDJSON":
            # sink_ndjson is streaming
            lazy_frame.sink_ndjson(path)

        elif fmt == "Excel":
            # Native Polars (Fast Eager Write)
            df = lazy_frame.collect()
            df.write_excel(path)

        elif fmt == "JSON":
            # Native Polars (Fast Eager Write)
            df = lazy_frame.collect()
            df.write_json(path)

        elif fmt == "SQLite":
            # SQLite Export (Eager)
            table = params.get('table', 'data') if isinstance(
                params, dict) else getattr(params, 'table', 'data')
            if_exists = params.get('if_exists', 'replace') if isinstance(
                params, dict) else getattr(params, 'if_exists', 'replace')
            valid_if_exists = cast(
                Literal['fail', 'replace', 'append'], if_exists)

            df = lazy_frame.collect()
            # Construct connection string
            # write_database supports "sqlite:///path.db"
            uri = f"sqlite:///{path}"
            df.write_database(table_name=table, connection=uri,
                              if_table_exists=valid_if_exists, engine="sqlalchemy")

        # --- FINAL METADATA ---
        size_str = "Unknown"
        if os.path.exists(path):
            size_bytes = os.path.getsize(path)
            size_mb = size_bytes / 1024 / 1024
            if size_mb < 1:
                size_str = f"{size_bytes / 1024:.2f} KB"
            else:
                size_str = f"{size_mb:.2f} MB"

        result_container['status'] = "Done"
        result_container['file_details'] = [{
            "name": os.path.basename(path),
            "path": path,
            "size": size_str
        }]
    except Exception as e:
        result_container['status'] = f"Error: {e}"
